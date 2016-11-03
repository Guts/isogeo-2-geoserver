# -*- coding: UTF-8 -*-
#!/usr/bin/env python
from __future__ import (absolute_import, print_function, unicode_literals)

# ----------------------------------------------------------------------------
# Name:         Infos from GeoServer
# Purpose:      Uses REST API to read GeoServer
#
# Author:       Julien Moura (https://github.com/Guts/)
#
# Python:       2.7.x
# Created:      04/04/2016
# Updated:      2016
# Licence:      GPL 3
# ----------------------------------------------------------------------------

# ############################################################################
# ########## Libraries #############
# ##################################
# Standard library
import logging
from logging.handlers import RotatingFileHandler
from os import path
from uuid import UUID

# Python 3 backported
from collections import OrderedDict

# 3rd party libraries
from geoserver.catalog import Catalog
from geoserver.resource import Coverage, FeatureType
from isogeo_pysdk import Isogeo
from openpyxl import load_workbook, Workbook
from openpyxl.cell import get_column_letter
from openpyxl.worksheet.properties import WorksheetProperties

# ############################################################################
# ########## GLOBALS ###############
# ##################################
# url_base = "https://preprod.ppige-npdc.fr"
# url_base = "https://www.ppige-npdc.fr"
# input_xlsx = r"Correspondance services - métadonnées v3.xlsx"

# LOG FILE ##
logger = logging.getLogger()
logging.captureWarnings(True)
logger.setLevel(logging.INFO)  # all errors will be get
log_form = logging.Formatter("%(asctime)s || %(levelname)s "
                             "|| %(module)s || %(message)s")
logfile = RotatingFileHandler("LOG_infos_geoserver.log", "a", 10000000, 2)
logfile.setLevel(logging.INFO)
logfile.setFormatter(log_form)
logger.addHandler(logfile)
logger.info('=================================================')

# ############################################################################
# ########## FUNCTIONS #############
# ##################################


def tunning_worksheets(li_worksheets):
    """CLEAN UP & TUNNING worksheets list."""
    for sheet in li_worksheets:
        # Freezing panes
        c_freezed = sheet['B2']
        sheet.freeze_panes = c_freezed

        # Print properties
        sheet.print_options.horizontalCentered = True
        sheet.print_options.verticalCentered = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

        # Others properties
        wsprops = sheet.sheet_properties
        wsprops.filterMode = True

        # enable filters
        sheet.auto_filter.ref = str("A1:{}{}")\
                                    .format(get_column_letter(sheet.max_column),
                                            sheet.max_row)
    pass

def is_uuid(uuid_string, version=4):
    """Si uuid_string est un code hex valide mais pas un uuid valid,
    UUID() va quand même le convertir en uuid valide. Pour se prémunir
    de ce problème, on check la version original (sans les tirets) avec
        # le code hex généré qui doivent être les mêmes.
    """
    try:
        uid = UUID(uuid_string, version=version)
        return uid.hex == uuid_string.replace('-', '')
    except ValueError:
        return False
    except TypeError:
        logger.error("uuid must be a string. Not: {} ({})".format(type(uuid_string),
                                                                       uuid_string))
        return False


# ############################################################################
# ########## Classes #############
# ################################


class ReadGeoServer():
    def __init__(self, gs_axx, dico_gs, tipo, txt=''):
        """Use OGR functions to extract basic informations about geoserver.

        gs_axx = tuple like {url of a geoserver, user, password)
        dico_gs = dictionary to store
        tipo = format
        text = dictionary of text in the selected language
        """
        # connection
        cat = Catalog(gs_axx[0], gs_axx[1], gs_axx[2],
                      disable_ssl_certificate_validation=gs_axx[3])
        # print(dir(cat))



        # -- WORKSPACES -------------------------------------------------------
        workspaces = cat.get_workspaces()
        for wk in workspaces:
            # print(wk.name, wk.enabled, wk.resource_type, wk.wmsstore_url)
            dico_gs[wk.name] = wk.href, {}
        # print(dir(wk))

        # -- STORES -----------------------------------------------------------
        # stores = cat.get_stores()
        # for st in stores:
        #     # print(st.name, st.enabled, st.href, st.resource_type)
        #     if hasattr(st, 'url'):
        #         url = st.url
        #     elif hasattr(st, 'resource_url'):
        #         url = st.resource_url
        #     else:
        #         print(dir(st))

        #     dico_gs.get(st.workspace.name)[1][st.name] = {"ds_type": st.type,
        #                                                   "ds_url": url}

        # print(dir(st))
        # print(st.url)
        # print(st.workspace)
        # print(dir(st.workspace))
        # print(st.workspace.name)

        # -- LAYERS -----------------------------------------------------------
        # resources_target = cat.get_resources(workspace='ayants-droits')
        layers = cat.get_layers()
        logging.info("{} layers found".format(len(layers)))
        dico_layers = OrderedDict()
        for layer in layers:
            # print(layer.resource_type)
            lyr_title = layer.resource.title
            lyr_name = layer.name
            lyr_wkspace = layer.resource._workspace.name
            if type(layer.resource) is Coverage:
                lyr_type = "coverage"
            elif type(layer.resource) is FeatureType:
                lyr_type = "vector"
            else:
                lyr_type = type(layer.resource)

            # a log handshake
            logging.info("{} | {} | {} | {}".format(layers.index(layer),
                                                    lyr_type,
                                                    lyr_name,
                                                    lyr_title))

            # # METADATA LINKS #
            # download link
            md_link_dl = "{0}/geoserver/{1}/ows?request=GetFeature"\
                         "&service=WFS&typeName={1}%3A{2}&version=2.0.0"\
                         "&outputFormat=SHAPE-ZIP"\
                         .format(url_base,
                                 lyr_wkspace,
                                 lyr_name)

            # mapfish links
            md_link_mapfish_wms = "{0}/mapfishapp/?layername={1}"\
                                  "&owstype=WMSLayer&owsurl={0}/"\
                                  "geoserver/{2}/ows"\
                                  .format(url_base,
                                          lyr_name,
                                          lyr_wkspace)

            md_link_mapfish_wfs = "{0}/mapfishapp/?layername={1}"\
                                  "&owstype=WFSLayer&owsurl={0}/"\
                                  "geoserver/{2}/ows"\
                                  .format(url_base,
                                          lyr_name,
                                          lyr_wkspace)

            md_link_mapfish_wcs = "{0}/mapfishapp/?cache=PreferNetwork"\
                                  "&crs=EPSG:2154&format=GeoTIFF"\
                                  "&identifier={1}:{2}"\
                                  "&url={0}/geoserver/ows?"\
                                  .format(url_base,
                                          lyr_wkspace,
                                          lyr_name)

            # OC links
            md_link_oc_wms = "{0}/geoserver/{1}/wms?layers={1}:{2}"\
                             .format(url_base,
                                     lyr_wkspace,
                                     lyr_name)
            md_link_oc_wfs = "{0}/geoserver/{1}/ows?typeName={1}:{2}"\
                             .format(url_base,
                                     lyr_wkspace,
                                     lyr_name)

            md_link_oc_wcs = "{0}/geoserver/{1}/ows?typeName={1}:{2}"\
                             .format(url_base,
                                     lyr_wkspace,
                                     lyr_name)

            # CSW Querier links
            md_link_csw_wms = "{0}/geoserver/ows?service=wms&version=1.3.0"\
                              "&request=GetCapabilities".format(url_base)

            md_link_csw_wfs = "{0}/geoserver/ows?service=wfs&version=2.0.0"\
                              "&request=GetCapabilities".format(url_base)

            # # # SERVICE LINKS #
            # GeoServer Edit links
            gs_link_edit = "{}/geoserver/web/?wicket:bookmarkablePage="\
                           ":org.geoserver.web.data.resource."\
                           "ResourceConfigurationPage"\
                           "&name={}"\
                           "&wsName={}".format(url_base,
                                               lyr_wkspace,
                                               lyr_name)

            # Metadata links (service => metadata)
            if is_uuid(dict_match_gs_md.get(lyr_name)):
                # HTML metadata
                md_uuid_pure = dict_match_gs_md.get(lyr_name)
                srv_link_html = "{}/portail/geocatalogue?uuid={}"\
                                .format(url_base, md_uuid_pure)

                # XML metadata
                md_uuid_formatted = "{}-{}-{}-{}-{}".format(md_uuid_pure[:8],
                                                            md_uuid_pure[8:12],
                                                            md_uuid_pure[12:16],
                                                            md_uuid_pure[16:20],
                                                            md_uuid_pure[20:])
                srv_link_xml = "http://services.api.isogeo.com/ows/s/"\
                               "{1}/{2}?"\
                               "service=CSW&version=2.0.2&request=GetRecordById"\
                               "&id=urn:isogeo:metadata:uuid:{0}&"\
                               "elementsetname=full&outputSchema="\
                               "http://www.isotc211.org/2005/gmd"\
                               .format(md_uuid_formatted,
                                       csw_share_id,
                                       csw_share_token)
                # add to GeoServer layer
                rzourc = cat.get_resource(lyr_name,
                                          store=layer.resource._store.name)
                rzourc.metadata_links = [('text/html', 'ISO19115:2003', srv_link_html),
                                         ('text/xml', 'ISO19115:2003', srv_link_xml),
                                         ('text/html', 'TC211', srv_link_html),
                                         ('text/xml', 'TC211', srv_link_xml)]
                # rzourc.metadata_links.append(('text/html', 'other', 'hohoho'))
                cat.save(rzourc)

            else:
                logging.info("Service without metadata: {} ({})".format(lyr_name,
                                                                        dict_match_gs_md.get(lyr_name)))
                pass

            dico_layers[layer.name] = {"title": lyr_title,
                                       "workspace": lyr_wkspace,
                                       "store_name": layer.resource._store.name,
                                       "store_type": layer.resource._store.type,
                                       "lyr_type": lyr_type,
                                       "md_link_dl": md_link_dl,
                                       "md_link_mapfish": md_link_mapfish_wms,
                                       "md_link_mapfish_wms": md_link_mapfish_wms,
                                       "md_link_mapfish_wfs": md_link_mapfish_wfs,
                                       "md_link_mapfish_wcs": md_link_mapfish_wcs,
                                       "md_link_oc_wms": md_link_oc_wms,
                                       "md_link_oc_wfs": md_link_oc_wfs,
                                       "md_link_oc_wcs": md_link_oc_wcs,
                                       "md_link_csw_wms": md_link_csw_wms,
                                       "md_link_csw_wfs": md_link_csw_wfs,
                                       "gs_link_edit": gs_link_edit,
                                       "srv_link_html": srv_link_html,
                                       "srv_link_xml": srv_link_xml,
                                       "md_id_matching": md_uuid_pure
                                       }

            # mem clean up
            # del dico_layer

        # print(dico_gs.get(layer.resource._workspace.name)[1][layer.resource._store.name])
        # print(dir(layer.resource))
        # print(dir(layer.resource.writers))
        # print(dir(layer.resource.writers.get("metadataLinks").func_dict))
        # print(layer.resource.metadata)
        # print(layer.resource.metadata_links)

        dico_gs["layers"] = dico_layers

# ############################################################################
# ##### Stand alone program ########
# ##################################

if __name__ == '__main__':
    u""" standalone execution for tests. Paths are relative considering a test
    within the official repository (https://github.com/Guts/DicoGIS/)"""
    # ------------ Specific imports ---------------------
    from ConfigParser import SafeConfigParser   # to manage options.ini

    # ------------ Settings from ini file ----------------
    # if not path.isfile(path.realpath(r"settings.ini")):
    #     logging.error("settings.ini file missing.")
    #     raise ValueError("settings.ini file missing.")
    # else:
    #     pass

    config = SafeConfigParser()
    config.read(r"settings.ini")

    settings = {s: dict(config.items(s)) for s in config.sections()}

    # Isogeo
    app_id = settings.get('isogeo').get('app_id')
    app_secret = settings.get('isogeo').get('app_secret')
    app_lang = settings.get('isogeo').get('app_lang')
    csw_share_id = settings.get('isogeo').get('csw_share_id')
    csw_share_token = settings.get('isogeo').get('csw_share_token')

    # GeoServer
    gs_url = settings.get('geoserver').get('gs_url')
    gs_user = settings.get('geoserver').get('gs_user')
    gs_pswd = settings.get('geoserver').get('gs_pswd')
    gs_ssl_off = settings.get('geoserver').get('gs_ssl_off')

    # Output
    out_prefix = settings.get('output').get('out_prefix')
    url_base = settings.get('output').get('url_base')

    # Input
    input_xlsx = settings.get('input').get('in_matching')
    # ------------------------------------------------------------------------

    # METADATA Links for GeoServer
    wb = load_workbook(filename=path.normpath(input_xlsx),
                       read_only=True,
                       guess_types=True,
                       data_only=True,
                       # use_iterators=True
                       )
    ws = wb.worksheets[0]  # first sheet

    dict_match_gs_md = {}
    for row in ws.iter_rows(row_offset=1):
        dict_match_gs_md[row[4].value] = row[6].value

    # ------------ GEOSERVER -------------------------------------------------
    # listing WFS
    li_geoservers = [(gs_url,
                      gs_user,
                      gs_pswd,
                      gs_ssl_off),
                     ]

    # recipient datas
    dico_gs = OrderedDict()

    # read WFS
    for gs in li_geoservers:
        dico_gs.clear()
        logging.info("\n{0}: ".format(gs))
        ReadGeoServer(gs,
                      dico_gs,
                      'GeoServer')

        # print(dico_gs)
        # print(dico_gs.keys())
        # print(dico_gs.get('ayants-droits')[1].get("BD_TOPO_2015_VOIES_FERREES_ET_AUTRES"))
        # print(dico_gs.get('ayants-droits')[1].get("bd_topo_reseau_routier_route_primaire"))
        # print(dico_gs.get('ayants-droits')[1].keys())
        # print(dico_gs.get('layers'))

    # ------------------------------------------------------------------------

    # ------------ ISOGEO ----------------------------------------------------
    # instanciating the class
    isogeo = Isogeo(client_id=app_id,
                    client_secret=app_secret,
                    lang=app_lang)
    token = isogeo.connect()

    search_results = isogeo.search(token)
    search_results = search_results.get('results')

    # ------------------------------------------------------------------------

    # ## EXCELs ############
    # -- WMS -------------------------------------------------------
    wb_gs_full = Workbook()
    dest_gs_full = '{}_gs_full.xlsx'.format(out_prefix)

    ws_gs_full = wb_gs_full.active
    ws_gs_full.title = "GEOSERVER - FULL"

    ws_gs_full["A1"] = "GS_WORKSPACE"
    ws_gs_full["B1"] = "GS_DATASTORE_NAME"
    ws_gs_full["C1"] = "GS_DATASTORE_TYPE"
    ws_gs_full["D1"] = "GS_SOURCE_TYPE"
    ws_gs_full["E1"] = "GS_NOM"
    ws_gs_full["F1"] = "GS_TITRE"
    ws_gs_full["G1"] = "MD_UUID"

    # -- WMS -------------------------------------------------------
    wb_wms = Workbook()
    dest_wms = '{}_wms_OC.xlsx'.format(out_prefix)

    ws_wms = wb_wms.active
    ws_wms.title = "WMS"

    ws_wms["A1"] = "TITRE GEOSERVER"
    ws_wms["B1"] = "INTITULE"
    ws_wms["C1"] = "TYPE"
    ws_wms["D1"] = "URL"
    ws_wms["E1"] = "ACTION"
    ws_wms["F1"] = "ASSOCIER_A"
    ws_wms["G1"] = "GS_DATASTORE_TYPE"
    ws_wms["H1"] = "GS_SOURCE_TYPE"
    ws_wms["I1"] = "GS_WORKSPACE"
    ws_wms["J1"] = "GS_DATASTORE_NAME"

    # -- WFS -------------------------------------------------------
    wb_wfs = Workbook()
    dest_wfs = '{}_wfs_OC.xlsx'.format(out_prefix)

    ws_wfs = wb_wfs.active
    ws_wfs.title = "WFS"

    ws_wfs["A1"] = "TITRE GEOSERVER"
    ws_wfs["B1"] = "INTITULE"
    ws_wfs["C1"] = "TYPE"
    ws_wfs["D1"] = "URL"
    ws_wfs["E1"] = "ACTION"
    ws_wfs["F1"] = "ASSOCIER_A"
    ws_wfs["G1"] = "GS_DATASTORE_TYPE"
    ws_wfs["H1"] = "GS_SOURCE_TYPE"
    ws_wfs["I1"] = "GS_WORKSPACE"
    ws_wfs["J1"] = "GS_DATASTORE_NAME"

    # -- DOWNLOAD ----------------------------------------------------
    wb_download = Workbook()
    dest_download = '{}_download_wfs.xlsx'.format(out_prefix)

    ws_dl = wb_download.active
    ws_dl.title = "DOWNLOAD"

    ws_dl["A1"] = "TITRE GEOSERVER"
    ws_dl["B1"] = "INTITULE"
    ws_dl["C1"] = "KIND"
    ws_dl["D1"] = "URL"
    ws_dl["E1"] = "ACTION"
    ws_dl["F1"] = "ASSOCIER_A"
    ws_dl["G1"] = "GS_DATASTORE_TYPE"
    ws_dl["H1"] = "GS_SOURCE_TYPE"
    ws_dl["I1"] = "GS_WORKSPACE"
    ws_dl["J1"] = "GS_DATASTORE_NAME"

    # -- MAPFISH APP -------------------------------------------------
    wb_mapfish = Workbook()
    dest_mapfish = '{}_mapfish.xlsx'.format(out_prefix)

    ws_fish_wms = wb_mapfish.active
    ws_fish_wms.title = "MAPFISH - WMS"

    ws_fish_wms["A1"] = "TITRE GEOSERVER"
    ws_fish_wms["B1"] = "INTITULE"
    ws_fish_wms["C1"] = "KIND"
    ws_fish_wms["D1"] = "URL"
    ws_fish_wms["E1"] = "ACTION"
    ws_fish_wms["F1"] = "ASSOCIER_A"
    ws_fish_wms["G1"] = "GS_DATASTORE_TYPE"
    ws_fish_wms["H1"] = "GS_SOURCE_TYPE"
    ws_fish_wms["I1"] = "GS_WORKSPACE"
    ws_fish_wms["J1"] = "GS_DATASTORE_NAME"

    ws_fish_wfs = wb_mapfish.create_sheet()
    ws_fish_wfs.title = "MAPFISH - WFS"

    ws_fish_wfs["A1"] = "TITRE GEOSERVER"
    ws_fish_wfs["B1"] = "INTITULE"
    ws_fish_wfs["C1"] = "KIND"
    ws_fish_wfs["D1"] = "URL"
    ws_fish_wfs["E1"] = "ACTION"
    ws_fish_wfs["F1"] = "ASSOCIER_A"
    ws_fish_wfs["G1"] = "GS_DATASTORE_TYPE"
    ws_fish_wfs["H1"] = "GS_SOURCE_TYPE"
    ws_fish_wfs["I1"] = "GS_WORKSPACE"
    ws_fish_wfs["J1"] = "GS_DATASTORE_NAME"

    # -- CSW QUERIER ------------------------------------------------
    wb_cswquerier = Workbook()
    dest_cswquerier = '{}_cswquerier.xlsx'.format(out_prefix)

    ws_csw_wms = wb_cswquerier.active
    ws_csw_wms.title = "CSW QUERIER - WMS"

    ws_csw_wms["A1"] = "TITRE GEOSERVER"
    ws_csw_wms["B1"] = "INTITULE"
    ws_csw_wms["C1"] = "KIND"
    ws_csw_wms["D1"] = "URL"
    ws_csw_wms["E1"] = "ACTION"
    ws_csw_wms["F1"] = "ASSOCIER_A"
    ws_csw_wms["G1"] = "GS_DATASTORE_TYPE"
    ws_csw_wms["H1"] = "GS_SOURCE_TYPE"
    ws_csw_wms["I1"] = "GS_WORKSPACE"
    ws_csw_wms["J1"] = "GS_DATASTORE_NAME"

    ws_csw_wfs = wb_cswquerier.create_sheet()
    ws_csw_wfs.title = "CSW QUERIER - WFS"

    ws_csw_wfs["A1"] = "TITRE GEOSERVER"
    ws_csw_wfs["B1"] = "INTITULE"
    ws_csw_wfs["C1"] = "KIND"
    ws_csw_wfs["D1"] = "URL"
    ws_csw_wfs["E1"] = "ACTION"
    ws_csw_wfs["F1"] = "ASSOCIER_A"
    ws_csw_wfs["G1"] = "GS_DATASTORE_TYPE"
    ws_csw_wfs["H1"] = "GS_SOURCE_TYPE"
    ws_csw_wfs["I1"] = "GS_WORKSPACE"
    ws_csw_wfs["J1"] = "GS_DATASTORE_NAME"

    # -- PARSING LAYERS AND STORING ------------------------------------------
    layers = dico_gs.get('layers')
    for lyr in layers.keys():
        layer = layers.get(lyr)
        row = layers.keys().index(lyr) + 2
        # title
        ws_gs_full["F{}".format(row)] = layer.get("title")
        ws_wms["A{}".format(row)] = layer.get("title")
        ws_wfs["A{}".format(row)] = layer.get("title")
        ws_dl["A{}".format(row)] = layer.get("title")
        ws_fish_wms["A{}".format(row)] = layer.get("title")
        ws_fish_wfs["A{}".format(row)] = layer.get("title")
        ws_csw_wms["A{}".format(row)] = layer.get("title")
        ws_csw_wfs["A{}".format(row)] = layer.get("title")

        # label
        tronqued_title = layer.get("title").rsplit(" -")[0]
        ws_gs_full["E{}".format(row)] = lyr
        ws_wms["B{}".format(row)] = "Couche WMS - {}".format(tronqued_title)
        ws_wfs["B{}".format(row)] = "Couche WFS - {}".format(tronqued_title)
        ws_dl["B{}".format(row)] = "Extraire - {}".format(tronqued_title)
        ws_fish_wms["B{}".format(row)] = "Visualiseur - {} (WMS)".format(tronqued_title)
        ws_fish_wfs["B{}".format(row)] = "Visualiseur - {} (WFS)".format(tronqued_title)
        ws_csw_wms["B{}".format(row)] = lyr
        ws_csw_wfs["B{}".format(row)] = lyr

        # link type
        ws_wms["C{}".format(row)] = "wms"
        ws_wfs["C{}".format(row)] = "wfs"
        ws_dl["C{}".format(row)] = "data"
        ws_fish_wms["C{}".format(row)] = "url"
        ws_fish_wfs["C{}".format(row)] = "url"
        ws_csw_wms["C{}".format(row)] = "wms"
        ws_csw_wfs["C{}".format(row)] = "wfs"

        # url
        ws_wms["D{}".format(row)] = layer.get("md_link_oc_wms")
        ws_wfs["D{}".format(row)] = layer.get("md_link_oc_wfs")
        ws_dl["D{}".format(row)] = layer.get("md_link_dl")
        ws_fish_wms["D{}".format(row)] = layer.get("md_link_mapfish_wms")
        ws_fish_wfs["D{}".format(row)] = layer.get("md_link_mapfish_wfs")
        ws_csw_wms["D{}".format(row)] = layer.get("md_link_csw_wms")
        ws_csw_wfs["D{}".format(row)] = layer.get("md_link_csw_wfs")

        # action
        ws_wms["E{}".format(row)] = "view"
        ws_wfs["E{}".format(row)] = "view"
        ws_dl["E{}".format(row)] = "download"
        ws_fish_wms["E{}".format(row)] = "view"
        ws_fish_wfs["E{}".format(row)] = "view"
        ws_csw_wms["E{}".format(row)] = "view"
        ws_csw_wfs["E{}".format(row)] = "[view,download]"

        # Isogeo metadata
        ws_gs_full["G{}".format(row)] = layer.get("md_id_matching")
        ws_wms["F{}".format(row)] = layer.get("md_id_matching")
        ws_wfs["F{}".format(row)] = layer.get("md_id_matching")
        ws_dl["F{}".format(row)] = layer.get("md_id_matching")
        ws_fish_wms["F{}".format(row)] = layer.get("md_id_matching")
        ws_fish_wfs["F{}".format(row)] = layer.get("md_id_matching")
        ws_csw_wms["F{}".format(row)] = layer.get("md_id_matching")
        ws_csw_wfs["F{}".format(row)] = layer.get("md_id_matching")

        # datastore type
        ws_gs_full["C{}".format(row)] = layer.get("store_type")
        ws_wms["G{}".format(row)] = layer.get("store_type")
        ws_wfs["G{}".format(row)] = layer.get("store_type")
        ws_dl["G{}".format(row)] = layer.get("store_type")
        ws_fish_wms["G{}".format(row)] = layer.get("store_type")
        ws_fish_wfs["G{}".format(row)] = layer.get("store_type")
        ws_csw_wms["G{}".format(row)] = layer.get("store_type")
        ws_csw_wfs["G{}".format(row)] = layer.get("store_type")

        # source type
        ws_gs_full["D{}".format(row)] = layer.get("lyr_type")
        ws_wms["H{}".format(row)] = layer.get("lyr_type")
        ws_wfs["H{}".format(row)] = layer.get("lyr_type")
        ws_dl["H{}".format(row)] = layer.get("lyr_type")
        ws_fish_wms["H{}".format(row)] = layer.get("lyr_type")
        ws_fish_wfs["H{}".format(row)] = layer.get("lyr_type")
        ws_csw_wms["H{}".format(row)] = layer.get("lyr_type")
        ws_csw_wfs["H{}".format(row)] = layer.get("lyr_type")

        # workspace
        ws_gs_full["A{}".format(row)] = layer.get("workspace")
        ws_wms["I{}".format(row)] = layer.get("workspace")
        ws_wfs["I{}".format(row)] = layer.get("workspace")
        ws_dl["I{}".format(row)] = layer.get("workspace")
        ws_fish_wms["I{}".format(row)] = layer.get("workspace")
        ws_fish_wfs["I{}".format(row)] = layer.get("workspace")
        ws_csw_wms["I{}".format(row)] = layer.get("workspace")
        ws_csw_wfs["I{}".format(row)] = layer.get("workspace")

        # datastore name
        ws_gs_full["B{}".format(row)] = layer.get("store_name")
        ws_wms["J{}".format(row)] = layer.get("store_name")
        ws_wfs["J{}".format(row)] = layer.get("store_name")
        ws_dl["J{}".format(row)] = layer.get("store_name")
        ws_fish_wms["J{}".format(row)] = layer.get("store_name")
        ws_fish_wfs["J{}".format(row)] = layer.get("store_name")
        ws_csw_wms["J{}".format(row)] = layer.get("store_name")
        ws_csw_wfs["J{}".format(row)] = layer.get("store_name")

    # -- SERVICES  METADATA ---------------------------------------------------
    wb_srvmd = Workbook()
    dest_srvmd = '{}_srv_md.xlsx'.format(out_prefix)

    ws_srvmd = wb_srvmd.active
    ws_srvmd.title = "GEOSERVER_METADATA"

    ws_srvmd["A1"] = "UUID_METADATA"
    ws_srvmd["B1"] = "ISOGEO_NOM"
    ws_srvmd["C1"] = "ISOGEO_TITRE"
    ws_srvmd["D1"] = "ISOGEO_CHEMIN"
    ws_srvmd["E1"] = "ISOGEO_RESUME"
    ws_srvmd["F1"] = "URL_HTML"
    ws_srvmd["G1"] = "URL_XML"

    # -- METADATA LINK for EXTERNAL  ------------------------------------------
    wb_external = Workbook()
    dest_external = '{}_md_external.xlsx'.format(out_prefix)

    ws_external = wb_external.active
    ws_external.title = "METADATA_DIRECT_LINK"

    ws_external["A1"] = "URL_HTML"
    ws_external["B1"] = "INTITULE"
    ws_external["C1"] = "KIND"
    ws_external["D1"] = "ACTION"
    ws_external["E1"] = "ASSOCIER_A"

    # -- PARSING METADATA AND STORING ----------------------------------------
    for md in search_results:
        row = search_results.index(md) + 2
        md_uuid_pure = md.get("_id")
        md_uuid_formatted = "{}-{}-{}-{}-{}".format(md_uuid_pure[:8],
                                                    md_uuid_pure[8:12],
                                                    md_uuid_pure[12:16],
                                                    md_uuid_pure[16:20],
                                                    md_uuid_pure[20:])

        # HTML metadata
        srv_link_html = "{}/portail/geocatalogue?uuid={}"\
                        .format(url_base, md_uuid_pure)

        # XML metadata
        srv_link_xml = "http://services.api.isogeo.com/ows/s/"\
                       "{1}/"\
                       "{2}?"\
                       "service=CSW&version=2.0.2&request=GetRecordById"\
                       "&id=urn:isogeo:metadata:uuid:{0}&"\
                       "elementsetname=full&outputSchema="\
                       "http://www.isotc211.org/2005/gmd"\
                       .format(md_uuid_formatted,
                               csw_share_id,
                               csw_share_token)

        # storing
        ws_srvmd["A{}".format(row)] = md_uuid_pure
        ws_srvmd["B{}".format(row)] = md.get("title", "")
        ws_srvmd["C{}".format(row)] = md.get("name", "")
        ws_srvmd["D{}".format(row)] = md.get("path", "")
        ws_srvmd["E{}".format(row)] = md.get("abstract", "")
        ws_srvmd["F{}".format(row)] = srv_link_html
        ws_srvmd["G{}".format(row)] = srv_link_xml

        ws_external["A{}".format(row)] = srv_link_html
        ws_external["B{}".format(row)] = "Voir la métadonnée originale - {}"\
                                         .format(md.get("title", ""))
        ws_external["C{}".format(row)] = "url"
        ws_external["D{}".format(row)] = "other"
        ws_external["E{}".format(row)] = md_uuid_pure

    # -- TUNNING -------------------------------------------------------
    li_worksheets = [ws_dl,
                     ws_csw_wfs,
                     ws_csw_wms,
                     ws_external,
                     ws_fish_wfs,
                     ws_fish_wms,
                     ws_gs_full,
                     ws_srvmd,
                     ws_wfs,
                     ws_wms,
                     ]
    tunning_worksheets(li_worksheets)

    # -- SAVE -------------------------------------------------------
    wb_gs_full.save(filename=dest_gs_full)
    wb_wms.save(filename=dest_wms)
    wb_wfs.save(filename=dest_wfs)
    wb_download.save(filename=dest_download)
    wb_mapfish.save(filename=dest_mapfish)
    wb_cswquerier.save(filename=dest_cswquerier)
    wb_srvmd.save(filename=dest_srvmd)
    wb_external.save(filename=dest_external)

    logging.info("XSLX GENERATED. OVER.")
